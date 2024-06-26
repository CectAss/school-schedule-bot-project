import openpyxl


class FromSheet:
    def __init__(self, day_of_week, class_name, lesson_number):
        self.day_of_week = day_of_week
        self.class_name = class_name
        self.lesson_number = lesson_number
        self.excel_table = None

    def load_excel_table(self):
        if self.excel_table is None:
            workbook = openpyxl.open('file.xlsx', read_only=True)
            sheet_names = workbook.sheetnames
            sheet2 = workbook[sheet_names[1]]
            sheet = workbook.active
            self.excel_table = sheet, sheet2

    def get_everything(self):
        self.load_excel_table()

        sheet, sheet2 = self.excel_table
        if self.teacher_y() is None:
            teacher_name = 'Не назначен учитель'
        else:
            teacher_name = sheet2[self.teacher_y()][1].value

        if (
                1 + self.lesson_number + self.day_of_week * 8 is None or self.lesson_x() is None) or teacher_name == 'Не назначен учитель':
            lesson_name = 'Нет урока'
        else:
            lesson_y = 1 + self.lesson_number + self.day_of_week * 8
            lesson_name = sheet[lesson_y][self.lesson_x()].value
        return [teacher_name, lesson_name, self.day_of_week, self.class_name, self.lesson_number]

    def lesson_x(self):
        self.load_excel_table()

        sheet, sheet2 = self.excel_table
        j = 2
        class_count = 0
        while sheet[1][j].value is not None and j < sheet.max_column - 1:
            j += 1
            class_count += 1
        for i in range(2, class_count + 2):
            if self.class_name == sheet[1][i].value:
                return i

    def teacher_y(self):  # day of week - число от 0 до 5, class_name - пример 10т, lesson number число от 1 до 8
        self.load_excel_table()

        sheet, sheet2 = self.excel_table
        j = 3
        teacher_count = 0
        while sheet2[j][1].value is not None and j < sheet.max_row - 1:
            j += 1
            teacher_count += 1
        for i in range(3, teacher_count + 3):
            if self.class_name == sheet2[i][1 + self.lesson_number + self.day_of_week * 8].value:
                return i
